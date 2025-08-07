import tkinter as tk
from tkinter import filedialog, messagebox
import tkinter.font
import vlc
import re
from datetime import datetime
import ollama
import time
import subprocess
import tempfile
import os
import openpyxl

class SubtitleTranslatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Subtitle Translator with Embedded Subtitles")
        self.root.geometry("1000x700")
        self.root.configure(bg="#f4f4f4")

        # Subtitle font scaling
        self.subtitle_font_size = 22  # Initial font size

        # VLC instance
        self.instance = vlc.Instance()
        self.player = self.instance.media_player_new()
        self.media = None

        # GUI elements
        self.video_frame = tk.Frame(self.root, bg="black")
        self.video_frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.video_frame, bg="black", highlightthickness=0)
        self.canvas.pack(fill=tk.BOTH, expand=True)
        self.canvas.update()

        # Subtitle overlay frame (centered over video, now with improved style)
        self.subtitle_overlay = tk.Frame(self.video_frame, bg="black", highlightthickness=0)
        self.subtitle_overlay.place(relx=0.5, rely=0.85, anchor="center", relwidth=0.9)
        self.subtitle_overlay.config(padx=20, pady=10)
        self.last_subtitle_text = None  # Track last subtitle to prevent flicker

        # Translation box
        self.translation_box = None

        # Bind spacebar and canvas click for resume
        self.root.bind('<space>', lambda e: self.resume_video())
        self.canvas.bind('<Button-1>', lambda e: self.resume_video())
        self.canvas.bind('<Double-1>', self.toggle_fullscreen)
        # Bind arrow keys for seeking
        self.root.bind('<Right>', lambda e: self.seek_relative(10))
        self.root.bind('<Left>', lambda e: self.seek_relative(-10))

        # Controls frame (now for player controls)
        self.player_controls_frame = tk.Frame(self.root, bg="#f4f4f4")
        self.player_controls_frame.pack(pady=5)

        self.play_pause_btn = tk.Button(self.player_controls_frame, text="Play", command=self.toggle_play_pause, font=("Segoe UI", 11), bg="#e0e0e0", relief="flat", padx=10, pady=5)
        self.play_pause_btn.grid(row=0, column=0, padx=5)

        self.stop_btn = tk.Button(self.player_controls_frame, text="Stop", command=self.stop_video, font=("Segoe UI", 11), bg="#e0e0e0", relief="flat", padx=10, pady=5)
        self.stop_btn.grid(row=0, column=1, padx=5)

        self.seek_var = tk.DoubleVar()
        self.seek_bar = tk.Scale(self.player_controls_frame, from_=0, to=100, orient=tk.HORIZONTAL, variable=self.seek_var, showvalue=0, length=400, command=self.seek_video, bg="#f4f4f4", highlightthickness=0)
        self.seek_bar.grid(row=0, column=2, padx=10)

        self.time_label = tk.Label(self.player_controls_frame, text="00:00 / 00:00", font=("Segoe UI", 10), bg="#f4f4f4")
        self.time_label.grid(row=0, column=3, padx=5)

        # Move load buttons to a separate frame for clarity
        self.load_controls_frame = tk.Frame(self.root, bg="#f4f4f4")
        self.load_controls_frame.pack(pady=5)
        self.load_button = tk.Button(self.load_controls_frame, text="Load Video", command=self.load_video, font=("Segoe UI", 11), bg="#e0e0e0", relief="flat", padx=10, pady=5)
        self.load_button.grid(row=0, column=0, padx=5)
        self.load_subs_button = tk.Button(self.load_controls_frame, text="Load Subtitles", command=self.load_subtitles, font=("Segoe UI", 11), bg="#e0e0e0", relief="flat", padx=10, pady=5)
        self.load_subs_button.grid(row=0, column=1, padx=5)

        self.status_label = tk.Label(self.root, text="", font=("Segoe UI", 11), bg="#f4f4f4", fg="#333")
        self.status_label.pack(pady=5)

        # Subtitle data
        self.subtitles = []
        self.current_subtitle_index = -1
        self.subtitle_tracks = []
        self.selected_track = None

        # VLC event manager
        self.event_manager = self.player.event_manager()
        self.event_manager.event_attach(vlc.EventType.MediaPlayerTimeChanged, self.update_subtitles)

        # Ollama model
        self.ollama_model = "gemma3:1b-it-qat"  # Change to your preferred model

        self.excel_file = 'translations.xlsx'
        self.setup_excel_file()

        # Bind configure event after all initializations
        self.root.bind('<Configure>', self.update_font_size)

    def toggle_play_pause(self):
        if self.player.is_playing():
            self.player.pause()
            self.play_pause_btn.config(text="Play")
        else:
            self.player.play()
            self.play_pause_btn.config(text="Pause")

    def stop_video(self):
        self.player.stop()
        self.play_pause_btn.config(text="Play")
        self.seek_var.set(0)
        self.time_label.config(text="00:00 / 00:00")

    def seek_video(self, value):
        if self.player.get_length() > 0:
            new_time = float(value) / 100 * self.player.get_length()
            self.player.set_time(int(new_time))

    def update_seek_bar(self):
        if self.player.get_length() > 0:
            pos = self.player.get_time() / self.player.get_length() * 100
            self.seek_var.set(pos)
            cur = self.format_time(self.player.get_time() // 1000)
            total = self.format_time(self.player.get_length() // 1000)
            self.time_label.config(text=f"{cur} / {total}")
        self.root.after(500, self.update_seek_bar)

    def format_time(self, seconds):
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        if h > 0:
            return f"{h:02}:{m:02}:{s:02}"
        else:
            return f"{m:02}:{s:02}"

    def load_video(self):
        video_file = filedialog.askopenfilename(filetypes=[("Video files", "*.mp4 *.mkv *.avi")])
        if video_file:
            self.media = self.instance.media_new(video_file)
            self.player.set_media(self.media)
            self.player.set_hwnd(self.canvas.winfo_id())
            self.player.play()
            self.status_label.config(text="Extracting embedded subtitles...")
            self.extract_embedded_subtitles(video_file)
            self.play_pause_btn.config(text="Pause")
            self.root.after(500, self.update_seek_bar)
        else:
            self.status_label.config(text="Please select a video file.")

    def load_subtitle_tracks(self):
        # Wait briefly for VLC to load media
        time.sleep(1)
        self.subtitle_tracks = []
        tracks = self.player.video_get_spu_description()
        if not tracks:
            messagebox.showerror("Error", "No embedded subtitle tracks found.")
            return

        for track in tracks:
            track_name = track[1].decode('utf-8', errors='ignore')
            track_id = track[0]
            self.subtitle_tracks.append((track_id, track_name))

        # Auto-select Spanish if available
        spanish_track = None
        for track_id, track_name in self.subtitle_tracks:
            if "spanish" in track_name.lower() or "español" in track_name.lower() or "es" in track_name.lower():
                spanish_track = (track_id, track_name)
                break

        if spanish_track:
            self.player.video_set_spu(spanish_track[0])
            self.selected_track = spanish_track
            self.status_label.config(text=f"Selected subtitle track: {spanish_track[1]}")
            self.extract_subtitles()
        else:
            self.select_subtitle_track()

    def select_subtitle_track(self):
        # Create a simple dialog to select subtitle track
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Subtitle Track")
        dialog.geometry("300x200")
        tk.Label(dialog, text="Select a subtitle track:").pack(pady=5)
        track_var = tk.StringVar()
        for track_id, track_name in self.subtitle_tracks:
            tk.Radiobutton(dialog, text=track_name, value=track_id, variable=track_var).pack(anchor="w")
        tk.Button(dialog, text="OK", command=lambda: self.set_subtitle_track(track_var.get(), dialog)).pack(pady=10)

    def set_subtitle_track(self, track_id, dialog):
        if track_id:
            track_id = int(track_id)
            self.player.video_set_spu(track_id)
            self.selected_track = next((t for t in self.subtitle_tracks if t[0] == track_id), None)
            self.status_label.config(text=f"Selected subtitle track: {self.selected_track[1]}")
            self.extract_subtitles()
        dialog.destroy()

    def extract_subtitles(self):
        # Note: python-vlc does not directly provide access to embedded subtitle text.
        # For simplicity, assume subtitles are rendered by VLC and we capture displayed text.
        # Advanced extraction requires external tools like ffmpeg to extract subtitle streams.
        self.status_label.config(text="Displaying embedded subtitles. Click words for translations.")
        # Since python-vlc can't extract subtitle text directly, we rely on VLC's rendering.
        # Subtitles are shown in the video, and we use a placeholder for clickable text.
        # For full subtitle text extraction, integrate ffmpeg or use VLC's subtitle display.

    def load_subtitles(self):
        srt_file = filedialog.askopenfilename(filetypes=[("Subtitle files", "*.srt")])
        if srt_file:
            self.subtitles = self.parse_srt_file(srt_file)
            self.status_label.config(text=f"Loaded subtitles: {srt_file}")
        else:
            self.status_label.config(text="No subtitle file selected.")

    def parse_srt_file(self, srt_path):
        with open(srt_path, 'r', encoding='utf-8') as f:
            content = f.read()
        pattern = re.compile(r'(\d+)\s+(\d{2}:\d{2}:\d{2},\d{3})\s*-->\s*(\d{2}:\d{2}:\d{2},\d{3})\s+(.+?)(?=\n\n|\Z)', re.DOTALL)
        subtitles = []
        for match in pattern.finditer(content):
            idx, start, end, text = match.groups()
            start_sec = self.srt_time_to_seconds(start)
            end_sec = self.srt_time_to_seconds(end)
            text = text.replace('\n', ' ').strip()
            subtitles.append({'start': start_sec, 'end': end_sec, 'text': text})
        return subtitles

    def srt_time_to_seconds(self, srt_time):
        h, m, s_ms = srt_time.split(':')
        s, ms = s_ms.split(',')
        return int(h) * 3600 + int(m) * 60 + int(s) + int(ms) / 1000

    def update_subtitles(self, event=None):
        self.root.update_idletasks() # Ensure root window geometry is updated
        current_time = self.player.get_time() / 1000  # VLC returns ms
        subtitle_line = ""
        for sub in self.subtitles:
            if sub['start'] <= current_time <= sub['end']:
                subtitle_line = sub['text']
                break
        if subtitle_line == self.last_subtitle_text:
            return  # No change, do not update (prevents flicker)
        self.last_subtitle_text = subtitle_line
        for widget in self.subtitle_overlay.winfo_children():
            widget.destroy()
        if subtitle_line:
            # Ensure the video frame is updated to get correct dimensions
            self.video_frame.update_idletasks()
            # Calculate the effective width for the subtitle overlay
            overlay_width = int(self.video_frame.winfo_width() * 0.9) - 2 * 20 # Subtract padx from subtitle_overlay.config

            self.safe_print(f"Video Frame Width: {self.video_frame.winfo_width()}")
            self.safe_print(f"Calculated Overlay Width for wrapping: {overlay_width}")

            current_line_frame = None
            current_line_width = 0
            words = subtitle_line.split()
            
            # Create a font object to measure text width
            font = tk.font.Font(family="Segoe UI", size=self.subtitle_font_size, weight="bold")

            for word in words:
                clean_word = re.sub(r"[.,!?]", "", word.lower())
                # Calculate word width including its own padx
                word_display_width = font.measure(word) + 2 * 4 # 2 * padx=4 for lbl.pack

                if current_line_frame is None or (current_line_width + word_display_width) > overlay_width:
                    # Start a new line
                    current_line_frame = tk.Frame(self.subtitle_overlay, bg="#222222", highlightthickness=0)
                    current_line_frame.pack(fill=tk.X, expand=True, pady=5) # Use fill=tk.X and expand=True
                    tk.Frame(current_line_frame, bg="#222222").pack(side=tk.LEFT, expand=True) # Left spacer
                    current_line_width = 0

                lbl = tk.Label(
                    current_line_frame,
                    text=word,
                    font=("Segoe UI", self.subtitle_font_size, "bold"),
                    bg="#222222",
                    fg="#fff",
                    bd=0,
                    relief="flat",
                    highlightthickness=0,
                    cursor="hand2"
                )
                lbl.pack(side=tk.LEFT, padx=4, pady=2)
                lbl.bind("<Button-1>", lambda e, w=clean_word, l=lbl: self.handle_word_click(w, l))
                current_line_width += word_display_width
            # Add right spacer after all words are processed for the last line
            tk.Frame(current_line_frame, bg="#222222").pack(side=tk.RIGHT, expand=True)
        else:
            self.safe_print("No subtitle to display at this time.")

    def handle_word_click(self, word, label_widget):
        self.safe_print(f"Word clicked: {word}")
        # Destroy existing translation box if it exists
        if self.translation_box:
            self.translation_box.destroy()
            self.translation_box = None

        # Cancel any pending hide job
        if self.translation_box_hide_job:
            self.root.after_cancel(self.translation_box_hide_job)
            self.translation_box_hide_job = None

        # Reset previous selected word's color if it exists and is still valid
        if self.selected_word_label and self.selected_word_label.winfo_exists():
            self.selected_word_label.config(bg="#222222")

        # Highlight the newly clicked word
        label_widget.config(bg="#ffe066")
        self.selected_word_label = label_widget

        translation = self.translate_word(word)
        self.show_translation_box(translation, label_widget)
        self.save_translation(word, translation, self.last_subtitle_text)

    def show_translation_box(self, translation, label_widget):
        self.safe_print(f"Showing translation box for: {translation}")
        # Create a new translation box
        self.translation_box = tk.Label(self.root, text=translation, font=("Segoe UI", 18, "italic"), bg="#fffbe6", fg="#333", bd=2, relief="ridge", padx=16, pady=8)

        # Place above the clicked word
        x = label_widget.winfo_rootx() - self.video_frame.winfo_rootx()
        y = label_widget.winfo_rooty() - self.video_frame.winfo_rooty() - self.translation_box.winfo_reqheight() - 25 # Adjusted higher
        self.safe_print(f"Placing translation box at x={x}, y={y}")
        self.translation_box.place(x=x, y=y)

        # Hide after 3 seconds
        self.translation_box_hide_job = self.root.after(3000, self.hide_translation_box)

    def hide_translation_box(self):
        if self.translation_box:
            self.translation_box.destroy()
            self.translation_box = None
        if self.selected_word_label and self.selected_word_label.winfo_exists():
            self.selected_word_label.config(bg="#222222")
            self.selected_word_label = None

    def translate_word(self, word):
        self.safe_print(f"Attempting to translate: {word}")
        try:
            prompt = f"Translate the Spanish word '{word}' to English. Provide only the translated word or a short phrase."
            response = ollama.generate(model=self.ollama_model, prompt=prompt)
            translated_text = response["response"].strip()
            self.safe_print(f"Translated '{word}' to: {translated_text}")
            return translated_text
        except Exception as e:
            self.safe_print(f"Error translating '{word}': {str(e)}")
            return f"Error translating: {str(e)}"

    def extract_embedded_subtitles(self, video_file):
        # Get subtitle tracks using ffmpeg
        try:
            cmd = [
                "ffmpeg", "-hide_banner", "-i", video_file
            ]
            result = subprocess.run(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            output = result.stderr
            # Find subtitle tracks
            import re
            tracks = re.findall(r'Stream #(\d+):(\d+)(?:\((\w*)\))?: Subtitle: ([^\n]+)', output)
            if not tracks:
                self.status_label.config(text="No embedded subtitle tracks found.")
                return
            # Prompt user to select track
            dialog = tk.Toplevel(self.root)
            dialog.title("Select Subtitle Track to Extract")
            dialog.geometry("500x300")
            tk.Label(dialog, text="Select a subtitle track to extract:", font=("Segoe UI", 11)).pack(pady=5)
            track_var = tk.IntVar(value=0)
            for idx, (stream_idx, track_idx, lang, desc) in enumerate(tracks):
                lang_str = f" ({lang})" if lang else ""
                tk.Radiobutton(dialog, text=f"Track {idx}: Stream {stream_idx}:{track_idx}{lang_str} - {desc}", value=idx, variable=track_var, font=("Segoe UI", 10)).pack(anchor="w")
            def on_ok():
                selected_idx = track_var.get()
                dialog.destroy()
                self.selected_subtitle_info = tracks[selected_idx]
                self.status_label.config(text=f"Using subtitle track: Stream {tracks[selected_idx][0]}:{tracks[selected_idx][1]} ({tracks[selected_idx][2]}) - {tracks[selected_idx][3]}")
                self.run_ffmpeg_extract(video_file, selected_idx)
            tk.Button(dialog, text="OK", command=on_ok, font=("Segoe UI", 11), bg="#e0e0e0").pack(pady=10)
        except Exception as e:
            self.status_label.config(text=f"Error extracting subtitle tracks: {e}")

    def run_ffmpeg_extract(self, video_file, track_idx):
        # Extract the selected subtitle track to a temp .srt file
        try:
            temp_dir = tempfile.gettempdir()
            srt_path = os.path.join(temp_dir, f"extracted_subs_{os.getpid()}.srt")
            cmd = [
                "ffmpeg", "-y", "-i", video_file, f"-map", f"0:s:{track_idx}", srt_path
            ]
            result = subprocess.run(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            if os.path.exists(srt_path):
                self.safe_print(f"Extracted SRT file: {srt_path}, size: {os.path.getsize(srt_path)} bytes")
            if os.path.exists(srt_path) and os.path.getsize(srt_path) > 0:
                self.subtitles = self.parse_srt_file(srt_path)
                self.safe_print(f"Loaded {len(self.subtitles)} subtitles from {srt_path}")
                if self.subtitles:
                    self.safe_print(f"First subtitle: {self.subtitles[0]}")
                    self.safe_print(f"Last subtitle: {self.subtitles[-1]}")
                self.status_label.config(text=f"Loaded extracted subtitles from track {track_idx}.")
            else:
                self.safe_print("Failed to extract subtitles or no subtitles found in selected track.")
                self.status_label.config(text="Failed to extract subtitles or no subtitles found in selected track.")
        except Exception as e:
            self.safe_print(f"Error extracting subtitles: {e}")
            self.status_label.config(text=f"Error extracting subtitles: {e}")

    def safe_print(self, text):
        try:
            print(text)
        except UnicodeEncodeError:
            print(text.encode('utf-8', 'ignore').decode('ascii', 'ignore'))

    def resume_video(self):
        if not self.player.is_playing():
            self.player.play()
            self.play_pause_btn.config(text="Pause")

    def seek_relative(self, seconds):
        cur_time = self.player.get_time() // 1000  # in seconds
        new_time = max(0, cur_time + seconds)
        self.player.set_time(int(new_time * 1000))

    def run(self):
        self.root.mainloop()

    def toggle_fullscreen(self, event=None):
        self.root.attributes("-fullscreen", not self.root.attributes("-fullscreen"))

    def update_font_size(self, event=None):
        width = self.root.winfo_width()
        self.subtitle_font_size = max(14, int(width / 60))
        # Re-render current subtitle to apply new font size immediately
        self.update_subtitles(None) # Pass None as event, as it's not a MediaPlayerTimeChanged event

    def setup_excel_file(self):
        if not os.path.exists(self.excel_file):
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Translations"
            sheet.append(["Original Word", "Translation", "Sentence"])
            workbook.save(self.excel_file)

    def save_translation(self, word, translation, sentence):
        workbook = openpyxl.load_workbook(self.excel_file)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == word:
                return # Word already exists
        sheet.append([word, translation, sentence])
        workbook.save(self.excel_file)

if __name__ == "__main__":
    root = tk.Tk()
    app = SubtitleTranslatorApp(root)
    app.run()