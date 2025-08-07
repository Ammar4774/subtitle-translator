import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import tkinter.font
import vlc
import re
import ollama
import time
import subprocess
import tempfile
import os
import openpyxl
from datetime import datetime
import threading
import webbrowser

class SubtitleTranslatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Enhanced Subtitle Translator")
        self.root.geometry("1000x700")
        self.root.configure(bg="#2c3e50")
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        # Subtitle font scaling
        self.subtitle_font_size = 22

        # VLC instance
        self.instance = vlc.Instance()
        self.player = self.instance.media_player_new()
        self.media = None
        self.video_paused = False
        self.fullscreen = False
        self.playback_speed = 1.0  # Default speed
        self.embedded_tracks = []  # Store detected subtitle tracks

        # GUI elements
        self.video_frame = tk.Frame(self.root, bg="black")
        self.video_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.video_frame.bind("<Double-Button-1>", self.toggle_fullscreen)

        # Video placeholder
        self.video_placeholder = tk.Label(
            self.video_frame, 
            text="Load a video file to begin", 
            font=("Segoe UI", 16), 
            bg="#34495e", 
            fg="#ecf0f1"
        )
        self.video_placeholder.pack(fill=tk.BOTH, expand=True)
        
        # Subtitle overlay
        self.subtitle_overlay = tk.Frame(self.video_frame, bg="black", highlightthickness=0)
        self.subtitle_overlay.place(relx=0.5, rely=0.85, anchor="center", relwidth=0.9)
        self.last_subtitle_text = None

        # Controls frame
        self.controls_frame = tk.Frame(self.root, bg="#2c3e50")
        self.controls_frame.pack(fill=tk.X, padx=10, pady=5)

        # Player controls
        self.player_controls = tk.Frame(self.controls_frame, bg="#2c3e50")
        self.player_controls.pack(side=tk.LEFT)

        self.play_pause_btn = self.create_button(
            self.player_controls, "▶", self.toggle_play_pause, "#27ae60", "#2ecc71"
        )
        self.play_pause_btn.pack(side=tk.LEFT, padx=2)

        self.stop_btn = self.create_button(
            self.player_controls, "■", self.stop_video, "#e74c3c", "#c0392b"
        )
        self.stop_btn.pack(side=tk.LEFT, padx=2)

        # Playback speed control
        self.speed_frame = tk.Frame(self.player_controls, bg="#2c3e50")
        self.speed_frame.pack(side=tk.LEFT, padx=5)
        
        tk.Label(
            self.speed_frame, text="Speed:", 
            font=("Segoe UI", 9), bg="#2c3e50", fg="#ecf0f1"
        ).pack(side=tk.LEFT)
        
        self.speed_var = tk.StringVar(value="1.0x")
        speed_options = ["0.5x", "0.75x", "1.0x", "1.25x", "1.5x", "2.0x"]
        self.speed_menu = ttk.Combobox(
            self.speed_frame, textvariable=self.speed_var, 
            values=speed_options, width=5, state="readonly"
        )
        self.speed_menu.pack(side=tk.LEFT)
        self.speed_menu.bind("<<ComboboxSelected>>", self.change_speed)

        # Seek bar
        self.seek_frame = tk.Frame(self.controls_frame, bg="#2c3e50")
        self.seek_frame.pack(fill=tk.X, expand=True, padx=10)

        self.seek_var = tk.DoubleVar()
        self.seek_bar = tk.Scale(
            self.seek_frame, from_=0, to=100, orient=tk.HORIZONTAL, 
            variable=self.seek_var, showvalue=0, length=400, 
            command=self.seek_video, bg="#34495e", sliderrelief="flat",
            troughcolor="#7f8c8d", highlightthickness=0
        )
        self.seek_bar.pack(fill=tk.X, expand=True, side=tk.LEFT)

        self.time_label = tk.Label(
            self.seek_frame, text="00:00 / 00:00", 
            font=("Segoe UI", 10), bg="#2c3e50", fg="#ecf0f1", width=12
        )
        self.time_label.pack(side=tk.LEFT, padx=5)

        # Load controls
        self.load_controls = tk.Frame(self.controls_frame, bg="#2c3e50")
        self.load_controls.pack(side=tk.RIGHT)

        self.load_button = self.create_button(
            self.load_controls, "Load Video", self.load_video, "#3498db", "#2980b9"
        )
        self.load_button.pack(side=tk.LEFT, padx=2)

        self.load_subs_button = self.create_button(
            self.load_controls, "Load Subtitles", self.load_subtitles, "#9b59b6", "#8e44ad"
        )
        self.load_subs_button.pack(side=tk.LEFT, padx=2)

        # Export button
        self.export_button = self.create_button(
            self.load_controls, "Export Translations", self.export_translations, "#f39c12", "#d35400"
        )
        self.export_button.pack(side=tk.LEFT, padx=2)

        # Toggle translation button
        self.toggle_trans_button = self.create_button(
            self.load_controls, "Hide Translations", self.toggle_translations, "#2ecc71", "#27ae60"
        )
        self.toggle_trans_button.pack(side=tk.LEFT, padx=2)

        # Status bar
        self.status_bar = tk.Label(
            self.root, text="Ready", font=("Segoe UI", 10), 
            bg="#34495e", fg="#ecf0f1", anchor=tk.W, padx=10
        )
        self.status_bar.pack(fill=tk.X, side=tk.BOTTOM, padx=10, pady=5)

        # Translation display
        self.translation_frame = tk.Frame(self.root, bg="#1a1a1a", height=60)
        self.translation_frame.pack(fill=tk.X, padx=10, pady=5)
        self.translation_label = tk.Label(
            self.translation_frame, text="", font=("Segoe UI", 14, "italic"), 
            bg="#1a1a1a", fg="#f39c12", wraplength=980, justify="center"
        )
        self.translation_label.pack(fill=tk.X, padx=10, pady=5)

        # Subtitle data
        self.subtitles = []
        self.current_subtitle_index = -1
        self.translation_cache = {}
        self.translations_visible = True

        # Ollama model
        self.ollama_model = "gemma3:1b-it-qat"
        self.excel_file = 'translations.xlsx'
        self.setup_excel_file()

        # Bindings
        self.root.bind('<space>', lambda e: self.toggle_play_pause())
        self.root.bind('<Escape>', self.exit_fullscreen)
        self.root.bind('<Right>', lambda e: self.seek_relative(5))
        self.root.bind('<Left>', lambda e: self.seek_relative(-5))
        self.root.bind('<Configure>', self.update_font_size)

    def log_debug(self, message):
        log_file_path = os.path.join(os.path.dirname(__file__), "debug_log.txt")
        with open(log_file_path, "a") as f:
            f.write(f"[{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}] {message}\n")

    def create_button(self, parent, text, command, bg, active_bg):
        return tk.Button(
            parent, text=text, command=command, font=("Segoe UI", 11),
            bg=bg, fg="white", activebackground=active_bg, activeforeground="white",
            relief="flat", padx=12, pady=5, bd=0, cursor="hand2"
        )

    def update_status(self, message):
        self.status_bar.config(text=message)
        self.root.update_idletasks()

    def load_video(self):
        video_file = filedialog.askopenfilename(
            filetypes=[("Video files", "*.mp4 *.mkv *.avi *.mov *.flv *.wmv")]
        )
        if video_file:
            try:
                self.video_placeholder.pack_forget()
                self.media = self.instance.media_new(video_file)
                self.player.set_media(self.media)
                
                # Get window handle after video frame is visible
                self.video_frame.update_idletasks()
                self.player.set_hwnd(self.video_frame.winfo_id())
                
                self.player.play()
                self.player.set_rate(self.playback_speed)
                self.play_pause_btn.config(text="⏸")
                self.video_paused = False
                self.root.after(500, self.update_seek_bar)
                self.update_status(f"Playing: {os.path.basename(video_file)}")
                
                # Check for embedded subtitles
                self.update_status("Checking for embedded subtitles...")
                threading.Thread(
                    target=self.check_embedded_subtitles, 
                    args=(video_file,), 
                    daemon=True
                ).start()
            except Exception as e:
                self.update_status(f"Error loading video: {str(e)}")
                messagebox.showerror("Error", f"Failed to load video:\n{str(e)}")

    def check_embedded_subtitles(self, video_file):
        """Check if the video has embedded subtitles"""
        try:
            cmd = ["ffprobe", "-v", "error", "-select_streams", "s", 
                   "-show_entries", "stream=index:stream_tags=language", 
                   "-of", "csv=p=0", video_file]
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            
            # Parse output for subtitle tracks
            tracks = []
            for line in result.stdout.splitlines():
                if line.strip():
                    parts = line.split(',')
                    if len(parts) >= 2:
                        index = parts[0].strip()
                        language = parts[1].strip() if len(parts) > 1 else "Unknown"
                        tracks.append((index, language))
            
            if tracks:
                self.embedded_tracks = tracks
                self.update_status(f"Found {len(tracks)} subtitle tracks")
                
                # Show track selection dialog in main thread
                self.root.after(0, lambda: self.show_track_selection(video_file))
            else:
                self.update_status("No embedded subtitles found. Please load external subtitles.")
        except Exception as e:
            self.update_status(f"Error checking subtitles: {str(e)}")

    def show_track_selection(self, video_file):
        """Show dialog to select subtitle track"""
        if not self.embedded_tracks:
            return
            
        dialog = tk.Toplevel(self.root)
        dialog.title("Select Subtitle Track")
        dialog.geometry("400x200")
        dialog.transient(self.root)
        dialog.grab_set()
        
        # Center dialog
        root_x = self.root.winfo_x()
        root_y = self.root.winfo_y()
        root_width = self.root.winfo_width()
        root_height = self.root.winfo_height()
        dialog_width = 400
        dialog_height = 200
        x = root_x + (root_width - dialog_width) // 2
        y = root_y + (root_height - dialog_height) // 2
        dialog.geometry(f"{dialog_width}x{dialog_height}+{x}+{y}")
        
        # Dialog content
        frame = tk.Frame(dialog, padx=20, pady=20, bg="#34495e")
        frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(frame, text="Select subtitle track:", 
                font=("Segoe UI", 12), bg="#34495e", fg="#ecf0f1").pack(pady=10)
        
        # Track selection combo
        track_var = tk.StringVar()
        track_options = [f"Track {idx} ({lang})" for idx, lang in self.embedded_tracks]
        track_combo = ttk.Combobox(
            frame, textvariable=track_var, 
            values=track_options, state="readonly", width=30
        )
        track_combo.current(0)
        track_combo.pack(pady=10)
        
        # Buttons
        btn_frame = tk.Frame(frame, bg="#34495e")
        btn_frame.pack(pady=10)
        
        def on_select():
            stream_index = selected_idx.get()
            srt_path = os.path.join(tempfile.gettempdir(), f"extracted_subs_{os.getpid()}.srt")

            cmd = [
                "ffmpeg", "-y", "-i", self.video_path,
                "-map", f"0:{stream_index}",
                "-c:s", "srt",  # Force SRT conversion
                srt_path
            ]

            result = subprocess.run(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            print("FFmpeg STDERR:\n", result.stderr)  # helpful for debugging

            if os.path.exists(srt_path) and os.path.getsize(srt_path) > 0:
                self.subtitles = parse_srt_file(srt_path)
                self.status_label.config(text=f"Loaded subtitle stream {stream_index}")
            else:
                self.status_label.config(text="Failed to extract selected subtitle track.")
            dialog.destroy()

        
        def on_cancel():
            dialog.destroy()
            self.update_status("Subtitle extraction canceled")
            
        tk.Button(btn_frame, text="Select", command=on_select, width=10,
                 bg="#27ae60", fg="white", relief="flat").pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Cancel", command=on_cancel, width=10,
                 bg="#e74c3c", fg="white", relief="flat").pack(side=tk.LEFT, padx=5)

    def extract_embedded_subtitles(self):
        if not self.video_path:
            messagebox.showwarning("Warning", "Please load a video first.")
            return

        try:
            # Use ffprobe to get subtitle streams
            probe_cmd = [
                "ffprobe", "-v", "error", "-select_streams", "s", "-show_entries",
                "stream=index:stream_tags=language", "-of", "csv=p=0", self.video_path
            ]
            probe_result = subprocess.run(probe_cmd, capture_output=True, text=True)
            lines = probe_result.stdout.strip().split("\n")
            tracks = [(line.split(',')[0], line.split(',')[1] if len(line.split(',')) > 1 else "Unknown")
                    for line in lines if line]

            if not tracks:
                self.status_label.config(text="No embedded subtitle tracks found.")
                return

            # Dialog to select track
            dialog = tk.Toplevel(self.root)
            dialog.title("Choose Subtitle Track")
            tk.Label(dialog, text="Select subtitle track to extract:").pack(pady=5)

            selected_idx = tk.StringVar(value=tracks[0][0])  # <-- moved to outer scope

            for index, lang in tracks:
                tk.Radiobutton(dialog, text=f"Stream {index} ({lang})", variable=selected_idx, value=index).pack(anchor="w")

        def on_select():
            stream_index = selected_idx.get()
            srt_path = os.path.join(tempfile.gettempdir(), f"extracted_subs_{os.getpid()}.srt")

            cmd = [
                "ffmpeg", "-y", "-i", self.video_path,
                "-map", f"0:{stream_index}",
                "-c:s", "srt",  # Force subtitle stream to convert to srt
                srt_path
            ]

            result = subprocess.run(cmd, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            print("FFmpeg stderr:\n", result.stderr)

            if os.path.exists(srt_path) and os.path.getsize(srt_path) > 0:
                self.subtitles = parse_srt_file(srt_path)
                self.status_label.config(text=f"Loaded subtitle stream {stream_index}")
            else:
                self.status_label.config(text="Failed to extract selected subtitle track.")
            dialog.destroy()

        tk.Button(dialog, text="Extract", command=on_select).pack(pady=10)

    except Exception as e:
        self.status_label.config(text=f"Subtitle extraction error: {e}")

    def parse_srt_file(self, srt_path):
        """Parse SRT subtitle files with improved error handling and multi-line support"""
        subtitles = []
        try:
            with open(srt_path, 'r', encoding='utf-8', errors='replace') as f:
                content = f.read()
                
            # Improved regex to handle various SRT formats and multi-line subtitles
            pattern = re.compile(
                r'(\d+)\s*?\n(\d{1,2}:\d{1,2}:\d{1,2}[,.]\d{1,3})\s*?-->\s*?(\d{1,2}:\d{1,2}:\d{1,2}[,.]\d{1,3})\s*?\n((?:.+\n)+?)\n'
            )
            
            for match in pattern.finditer(content):
                idx, start, end, text = match.groups()
                start_sec = self.srt_time_to_seconds(start.replace(',', '.'))
                end_sec = self.srt_time_to_seconds(end.replace(',', '.'))
                
                # Clean and process text
                text = re.sub(r'<[^>]+>', '', text)  # Remove HTML tags
                text = re.sub(r'\{[^}]+\}', '', text)  # Remove SSA tags
                text = re.sub(r'\n', ' ', text)  # Replace newlines with spaces
                text = re.sub(r'\s+', ' ', text).strip()  # Collapse whitespace
                
                subtitles.append({
                    'start': start_sec, 
                    'end': end_sec, 
                    'text': text,
                    'original': text
                })
                
            return subtitles
        except Exception as e:
            self.update_status(f"Error parsing subtitles: {str(e)}")
            return []

    def srt_time_to_seconds(self, srt_time):
        """Convert SRT time format to seconds"""
        parts = re.split(r'[:,.]', srt_time)
        if len(parts) == 4:  # HH:MM:SS.mmm
            h, m, s, ms = map(int, parts)
            return h * 3600 + m * 60 + s + ms / 1000.0
        elif len(parts) == 3:  # MM:SS.mmm
            m, s, ms = map(int, parts)
            return m * 60 + s + ms / 1000.0
        else:  # SS.mmm
            return float(srt_time)

    def update_subtitles(self, event=None):
        """Check and update visible subtitles periodically"""
        try:
            if not self.subtitles or not self.player.is_playing():
                self.root.after(100, self.update_subtitles)
                return
                
            current_time = self.player.get_time() / 1000.0  # ms to seconds
            current_sub = None
            self.log_debug(f"update_subtitles - Current time: {current_time:.2f}s")
            
            # Find current subtitle
            for i, sub in enumerate(self.subtitles):
                if sub['start'] <= current_time <= sub['end']:
                    current_sub = sub
                    self.current_subtitle_index = i
                    self.log_debug(f"Found current subtitle: {current_sub['text'][:50]}...")
                    break
                    
            # Clear if no subtitle
            if not current_sub:
                self.clear_subtitles()
                self.root.after(100, self.update_subtitles)
                return
                
            # Skip if same as last
            if current_sub['text'] == self.last_subtitle_text:
                self.root.after(100, self.update_subtitles)
                return
                
            self.last_subtitle_text = current_sub['text']
            self.render_subtitles(current_sub['text'])
            self.root.after(100, self.update_subtitles)
        except Exception as e:
            self.update_status(f"Subtitle error: {str(e)}")
            self.root.after(100, self.update_subtitles)

    def render_subtitles(self, text):
        """Render subtitles with clickable words"""
        self.log_debug(f"render_subtitles called with text: {text[:50]}...")
        # Clear existing subtitles
        for widget in self.subtitle_overlay.winfo_children():
            widget.destroy()
            
        # Create frame for current line
        line_frame = tk.Frame(self.subtitle_overlay, bg="#000000")
        line_frame.pack(fill=tk.X, pady=2)
        
        # Split text into words
        words = re.findall(r"[\w']+|[.,!?;]", text)
        
        # Add words as clickable labels
        for word in words:
            clean_word = re.sub(r"[.,!?;]", "", word.lower())
            lbl = tk.Label(
                line_frame, text=word, 
                font=("Segoe UI", self.subtitle_font_size, "bold"),
                bg="#000000", fg="#FFFFFF", cursor="hand2", padx=2
            )
            lbl.pack(side=tk.LEFT)
            lbl.bind("<Button-1>", lambda e, w=clean_word: self.handle_word_click(w))
            lbl.bind("<Enter>", lambda e, l=lbl: l.config(bg="#333333"))
            lbl.bind("<Leave>", lambda e, l=lbl: l.config(bg="#000000"))

    def clear_subtitles(self):
        """Clear subtitle display"""
        for widget in self.subtitle_overlay.winfo_children():
            widget.destroy()
        self.last_subtitle_text = None

    def handle_word_click(self, word):
        """Handle word click for translation"""
        if not self.video_paused:
            self.player.pause()
            self.play_pause_btn.config(text="▶")
            self.video_paused = True
            
        # Check cache first
        if word in self.translation_cache:
            translation = self.translation_cache[word]
        else:
            translation = self.translate_word(word)
            self.translation_cache[word] = translation
            
        self.show_translation(translation)
        self.save_translation(word, translation)

    def translate_word(self, word):
        """Translate word using Ollama"""
        if not word.strip():
            return "No word selected"
            
        self.update_status(f"Translating: {word}...")
        try:
            prompt = f"Translate the Spanish word '{word}' to English. Provide only the translated word or a short phrase."
            response = ollama.generate(model=self.ollama_model, prompt=prompt)
            return response["response"].strip()
        except Exception as e:
            return f"Translation error: {str(e)}"

    def show_translation(self, text):
        """Show translation in the translation box"""
        if self.translations_visible:
            self.translation_frame.pack(fill=tk.X, padx=10, pady=5)
            self.translation_label.config(text=text)
            self.root.after(3000, self.hide_translation)

    def hide_translation(self):
        """Hide translation box after timeout"""
        if self.translations_visible:
            self.translation_frame.pack_forget()

    def toggle_translations(self):
        """Toggle visibility of translation bar"""
        self.translations_visible = not self.translations_visible
        if self.translations_visible:
            self.toggle_trans_button.config(text="Hide Translations")
            self.translation_frame.pack(fill=tk.X, padx=10, pady=5)
        else:
            self.toggle_trans_button.config(text="Show Translations")
            self.translation_frame.pack_forget()

    def toggle_play_pause(self):
        """Toggle play/pause state"""
        if self.player.is_playing():
            self.player.pause()
            self.play_pause_btn.config(text="▶")
            self.video_paused = True
        else:
            self.player.play()
            self.player.set_rate(self.playback_speed)
            self.play_pause_btn.config(text="⏸")
            self.video_paused = False

    def stop_video(self):
        """Stop video playback"""
        self.player.stop()
        self.play_pause_btn.config(text="▶")
        self.seek_var.set(0)
        self.time_label.config(text="00:00 / 00:00")
        self.clear_subtitles()
        self.video_placeholder.pack(fill=tk.BOTH, expand=True)

    def update_seek_bar(self):
        """Update the seek bar position"""
        if self.player and self.player.get_media():
            length = self.player.get_length()
            if length > 0:
                pos = min(100, max(0, self.player.get_time() / length * 100))
                self.seek_var.set(pos)
                cur = self.format_time(self.player.get_time() // 1000)
                total = self.format_time(length // 1000)
                self.time_label.config(text=f"{cur} / {total}")
        self.root.after(500, self.update_seek_bar)

    def format_time(self, seconds):
        """Format time as HH:MM:SS or MM:SS"""
        m, s = divmod(int(seconds), 60)
        h, m = divmod(m, 60)
        return f"{h:02d}:{m:02d}:{s:02d}" if h else f"{m:02d}:{s:02d}"

    def seek_video(self, value):
        """Seek to position in video"""
        if self.player.get_media() and self.player.get_length() > 0:
            new_time = float(value) / 100 * self.player.get_length()
            self.player.set_time(int(new_time))

    def seek_relative(self, seconds):
        """Seek forward or backward by specified seconds"""
        if self.player.get_media():
            cur_time = self.player.get_time() // 1000
            new_time = max(0, cur_time + seconds)
            self.player.set_time(int(new_time * 1000))

    def change_speed(self, event):
        """Change playback speed"""
        speed_str = self.speed_var.get()
        speed_map = {
            "0.5x": 0.5,
            "0.75x": 0.75,
            "1.0x": 1.0,
            "1.25x": 1.25,
            "1.5x": 1.5,
            "2.0x": 2.0
        }
        self.playback_speed = speed_map.get(speed_str, 1.0)
        
        if self.player:
            self.player.set_rate(self.playback_speed)
            self.update_status(f"Playback speed: {speed_str}")

    def load_subtitles(self):
        """Load external subtitle file"""
        filetypes = [("Subtitle files", "*.srt *.ass *.vtt *.txt")]
        srt_file = filedialog.askopenfilename(filetypes=filetypes)
        if srt_file:
            self.subtitles = self.parse_srt_file(srt_file)
            if self.subtitles:
                self.update_status(f"Loaded {len(self.subtitles)} subtitles from {os.path.basename(srt_file)}")
                self.log_debug(f"Parsed {len(self.subtitles)} subtitles from external file: {os.path.basename(srt_file)}")
                # Start subtitle update loop
                self.root.after(100, self.update_subtitles)
            else:
                self.update_status("Failed to load subtitles")
                self.log_debug("Failed to load external subtitles.")

    def update_font_size(self, event=None):
        """Adjust subtitle font size based on window size"""
        if not self.fullscreen:
            width = self.root.winfo_width()
            self.subtitle_font_size = max(16, min(28, int(width / 45)))

    def setup_excel_file(self):
        """Create Excel file for translations if needed"""
        try:
            if not os.path.exists(self.excel_file):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Translations"
                ws.append(["Spanish", "English", "Timestamp"])
                wb.save(self.excel_file)
        except Exception as e:
            self.update_status(f"Excel error: {str(e)}")

    def save_translation(self, spanish, english):
        """Save translation to Excel file"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            ws = wb.active
            
            # Check for duplicates
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if row[0] == spanish:
                    return
                    
            # Add new entry
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([spanish, english, timestamp])
            wb.save(self.excel_file)
        except Exception as e:
            self.update_status(f"Save error: {str(e)}")

    def export_translations(self):
        """Open the translations Excel file"""
        try:
            if os.path.exists(self.excel_file):
                webbrowser.open(self.excel_file)
                self.update_status(f"Opened translations file: {self.excel_file}")
            else:
                self.update_status("No translations file found")
        except Exception as e:
            self.update_status(f"Error opening file: {str(e)}")

    def toggle_fullscreen(self, event=None):
        """Toggle fullscreen mode"""
        self.fullscreen = not self.fullscreen
        self.root.attributes("-fullscreen", self.fullscreen)
        if self.fullscreen:
            self.update_status("Fullscreen mode (Press ESC to exit)")
        else:
            self.update_status("Windowed mode")

    def exit_fullscreen(self, event=None):
        """Exit fullscreen mode"""
        if self.fullscreen:
            self.root.attributes("-fullscreen", False)
            self.fullscreen = False
            self.update_status("Exited fullscreen mode")

    def on_closing(self):
        """Handle window closing"""
        if self.player:
            self.player.stop()
        self.root.destroy()

    def run(self):
        """Start the application"""
        self.root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    app = SubtitleTranslatorApp(root)
    app.run()